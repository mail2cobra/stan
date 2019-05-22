# -*- coding: utf-8 -*-

"""
info:处理某数据文件，将word中的表格解析，保存数据到excel。
author:stan wong
github:
update_time:2019/05/21
依赖库：python-docx, openpyxl
"""

# 秘密文件，文件信息与内容不能外泄。

import os

from docx import Document
from openpyxl import Workbook


class Doc2Xls(object):
    def __init__(self, doc, xls):
        self.docFile = doc
        self.xlsFile = xls

    def SaveTable(self):
        print("开始处理文件...")

        doc = Document(self.docFile)
        tables = doc.tables

        table = tables[0]

        # 表头
        head = []
        head.append(table.cell(0, 0).text)
        head.append(table.cell(0, 2).text)
        head.append(table.cell(1, 0).text)
        head.append(table.cell(1, 2).text)
        head.append(table.cell(2, 0).text)
        head.append(table.cell(2, 2).text)
        head.append(table.cell(3, 1).text)
        head.append(table.cell(4, 1).text)
        head.append(table.cell(5, 0).text)
        head.append(table.cell(6, 0).text

        # 写入excel表头
        wb = Workbook()
        ws = wb.active
        ws.append(head)

        content = []
        row = 0
        for table in tables:
            print("第" + str(row) + "个:" + table.cell(0, 1).text)
            content.clear()
            content.append(table.cell(0, 1).text)
            content.append(table.cell(0, 3).text)
            content.append(table.cell(1, 1).text)
            content.append(table.cell(1, 3).text)
            content.append(table.cell(2, 1).text)
            content.append(table.cell(2, 3).text)
            content.append(table.cell(3, 2).text)
            content.append(table.cell(4, 2).text)
            content.append(table.cell(5, 1).text)
            t61 = table.cell(6, 1).text            # 内容
            try:
                t61 = t61 + table.cell(7, 1).text  # 如果有第七行的话，追加内容
            except Exception:
                pass
            content.append(t61)

            # 处理文本内容中的回车换行
            for i in range(0, len(content)):
                content[i] = str(content[i]).replace("\n", "")

            ws.append(content)
            row = row + 1

        wb.save(xls)
        print("处理完成：共保存了" + str(row) + "个表格。")


if __name__ == '__main__':

    print("word表格转存excel")
    print("----------------------------")

    while True:
        doc = input("+请输入要保存的word文件地址（d:\\test.doc）")
        if doc == "":
            print("提示：请重新输入。")
            continue

        if not os.path.exists(doc):
            print("提示：文件不存在，请重新输入。")
            continue

        xls = os.path.splitext(doc)[0] + ".xls"
        if os.path.exists(xls):
            print("提示：在doc文件目录下已经存在同名的xls文件，将覆盖该xls文件重新生成。")
            menu = input("是否确认(1-确认，0-取消)？")
            if menu != "1":
                print("错误：" + xls + "已经存在，请检查。")
                break

        d2x = Doc2Xls(doc, xls)
        d2x.SaveTable()
        break
 
